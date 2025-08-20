import tkinter as tk
from tkinter import filedialog, ttk, messagebox
try:
    import pdfplumber
except ModuleNotFoundError:  # pragma: no cover - handled at runtime
    pdfplumber = None
import re
import json
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
import subprocess
import sys
from style_utils import ABB_COLORS, aplicar_colorimetria

CONFIG_FILE = "column_config.json"
excel_path = None

DEFAULT_COLUMNS = {
    "Catalog Number": "A",
    "Power (HP)": "B",
    "Speed (RPM)": "C",
    "Phase": "D",
    "Hertz": "E",
    "Voltage": "F",
    "Order Codes": "G",
}


def load_column_config():
    try:
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)

        loaded_excel_path = data.get("excel_path")
        sheet_name = data.get("_sheet")
        columns = {**DEFAULT_COLUMNS, **data.get("columns", {})}

        if loaded_excel_path and os.path.isfile(loaded_excel_path):
            try:
                wb = load_workbook(loaded_excel_path)
                wb.close()
            except Exception:
                loaded_excel_path = None
        else:
            loaded_excel_path = None

        config = {"_sheet": sheet_name, **columns}
        return config, loaded_excel_path
    except Exception:
        return {"_sheet": None, **DEFAULT_COLUMNS.copy()}, None


column_config, excel_path = load_column_config()

def extraer_datos(pdf_path):
    datos = {
        "Catalog Number": None,
        "Power (HP)": None,
        "Speed (RPM)": None,
        "Phase": None,
        "Hertz": None,
        "Voltage": None,
        "Order Codes": []
    }

    if pdfplumber is None:
        messagebox.showerror(
            "Dependencia faltante",
            "El módulo pdfplumber no está instalado. "
            "Ejecuta 'pip install pdfplumber' e inténtalo de nuevo.",
        )
        return datos

    with pdfplumber.open(pdf_path) as pdf:
        text = ""
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text.replace("\n", " ") + " "

    # --- Catalog Number ---
    match_catalog = re.search(r"Catalog\s+Number\s+([A-Z0-9\-]+)", text, re.IGNORECASE)
    if match_catalog:
        datos["Catalog Number"] = match_catalog.group(1)

    # --- POWER ---
    match_power = re.search(r"Power\s+([\d.]+)\s*HP", text, re.IGNORECASE)
    if not match_power:
        match_power = re.search(r"\b([\d.]+)\s*HP\b", text)
    if match_power:
        datos["Power (HP)"] = match_power.group(1)

    # --- SPEED ---
    match_speed = re.search(r"Speed\s*\(RPM\)\s*(\d+)", text, re.IGNORECASE)
    if not match_speed:
        match_speed = re.search(r"\b(\d{3,4})\b\s*(?:RPM|R\.P\.M\.?)", text, re.IGNORECASE)
    if not match_speed:
        match_speed = re.search(r"\b(900|1200|1500|1800|3000|3600)\b", text)
    if match_speed:
        datos["Speed (RPM)"] = match_speed.group(1)

    # --- PHASE / HERTZ / VOLTAGE ---
    phv_matches = re.findall(r"\b(\d*)\s*/\s*(\d*)\s*/\s*([\d/]+)\b", text)
    for phase_val, hertz_val, volt_val in phv_matches:
        if len(volt_val) == 4 and volt_val.startswith("20"):  # evitar fechas
            continue
        datos["Phase"] = phase_val if phase_val else None
        datos["Hertz"] = hertz_val if hertz_val else None
        datos["Voltage"] = volt_val if volt_val else None
        break

    # --- ORDER CODES ---
    order_codes = re.findall(r"NEMA Motor Modifications\s+([A-Z0-9]{2,3})\s*-", text)
    datos["Order Codes"] = list(dict.fromkeys(order_codes))  # quitar duplicados manteniendo orden

    return datos

def save_column_config():
    try:
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            data = {
                "excel_path": excel_path,
                "_sheet": column_config.get("_sheet"),
                "columns": {k: v for k, v in column_config.items() if k != "_sheet"},
            }
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def configurar_columnas():
    if excel_path is None:
        messagebox.showwarning(
            "Excel no seleccionado", "Por favor selecciona un archivo Excel"
        )
        return

    # Cargar workbook y mapear hojas a columnas disponibles
    try:
        wb = load_workbook(excel_path)
        sheet_map = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            opciones = {}
            first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            if first_row:
                for idx, header in enumerate(first_row[0], start=1):
                    letra = get_column_letter(idx)
                    display = str(header).strip() if header else letra
                    opciones[display] = letra
            else:
                for idx in range(1, ws.max_column + 1):
                    letra = get_column_letter(idx)
                    opciones[letra] = letra
            if not opciones:
                for idx in range(1, 27):
                    letra = get_column_letter(idx)
                    opciones[letra] = letra
            sheet_map[sheet_name] = opciones
    except Exception as exc:
        messagebox.showerror("Error al leer Excel", str(exc))
        return
    finally:
        wb.close()

    config_win = tk.Toplevel(root)
    config_win.title("Configurar Columnas")
    aplicar_colorimetria(config_win)

    # Combobox para seleccionar hoja
    tk.Label(config_win, text="Hoja").grid(row=0, column=0, padx=5, pady=5, sticky="e")
    sheet_cb = ttk.Combobox(config_win, values=list(sheet_map.keys()), state="readonly")
    hoja_inicial = column_config.get("_sheet")
    if hoja_inicial not in sheet_map:
        hoja_inicial = list(sheet_map.keys())[0]
    sheet_cb.set(hoja_inicial)
    sheet_cb.grid(row=0, column=1, padx=5, pady=5)

    # Combobox para columnas por campo
    field_names = [k for k in column_config.keys() if k != "_sheet"]
    combos = {}
    for i, campo in enumerate(field_names):
        tk.Label(config_win, text=campo).grid(
            row=i + 1, column=0, padx=5, pady=5, sticky="e"
        )
        cb = ttk.Combobox(config_win, state="readonly")
        cb.grid(row=i + 1, column=1, padx=5, pady=5)
        combos[campo] = cb

    def cargar_opciones(sheet_name):
        opciones = list(sheet_map[sheet_name].keys())
        for campo, cb in combos.items():
            cb["values"] = opciones
            letra = column_config.get(campo, DEFAULT_COLUMNS[campo])
            display = next(
                (d for d, l in sheet_map[sheet_name].items() if l == letra),
                opciones[0] if opciones else "",
            )
            cb.set(display)

    cargar_opciones(hoja_inicial)
    sheet_cb.bind("<<ComboboxSelected>>", lambda e: cargar_opciones(sheet_cb.get()))

    def guardar():
        column_config["_sheet"] = sheet_cb.get()
        opciones = sheet_map[column_config["_sheet"]]
        for campo, cb in combos.items():
            seleccion = cb.get()
            columna = opciones.get(seleccion, seleccion)
            column_config[campo] = columna.upper() or DEFAULT_COLUMNS[campo]
        save_column_config()
        config_win.destroy()

    ttk.Button(config_win, text="Guardar", command=guardar).grid(
        row=len(combos) + 1, columnspan=2, pady=10
    )


def seleccionar_excel():
    global excel_path
    excel_path = filedialog.askopenfilename(
        title="Selecciona un archivo Excel",
        filetypes=[("Archivos Excel", "*.xlsx")]
    )
    if excel_path:
        try:
            wb = load_workbook(excel_path)
            wb.close()
        except Exception as exc:
            messagebox.showerror("Error al abrir Excel", str(exc))
            excel_path = None
        save_column_config()


def find_next_empty_row(sheet, columns):
    """Return the first row index where all given columns are empty.

    Recorre desde la fila 2 (asumiendo encabezados en la fila 1) y busca la
    primera fila en la que todas las columnas especificadas estén vacías. Si no
    encuentra ninguna, devuelve la fila siguiente a `sheet.max_row`.
    """

    for row in range(2, sheet.max_row + 1):
        if all(sheet[f"{col}{row}"].value in (None, "") for col in columns):
            return row
    return sheet.max_row + 1


def guardar_en_excel(datos):
    if excel_path is None:
        messagebox.showwarning(
            "Excel no seleccionado", "Por favor selecciona un archivo Excel"
        )
        return

    wb = load_workbook(excel_path)
    sheet_name = column_config.get("_sheet")
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    columnas = [col for campo, col in column_config.items() if campo != "_sheet"]
    row = find_next_empty_row(ws, columnas)
    for campo, columna in column_config.items():
        if campo == "_sheet":
            continue
        valor = datos.get(campo)
        if campo == "Order Codes":
            valor = ", ".join(valor)

        if isinstance(valor, str) and valor.replace(".", "", 1).isdigit():
            if "." in valor:
                valor = float(valor)
            else:
                valor = int(valor)

        ws[f"{columna}{row}"] = valor
    wb.save(excel_path)
    wb.close()


def ejecutar_sap():
    """Ejecuta el script que descarga documentos desde SAP."""
    try:
        subprocess.run([sys.executable, "sap_script.py"], check=True)
        messagebox.showinfo("SAP", "Descarga completada")
    except Exception as exc:
        messagebox.showerror("Error SAP", str(exc))


def seleccionar_pdf():
    file_path = filedialog.askopenfilename(
        title="Selecciona un archivo PDF",
        filetypes=[("Archivos PDF", "*.pdf")]
    )
    if file_path:
        datos = extraer_datos(file_path)
        mostrar_datos(datos)
        guardar_en_excel(datos)
        print(datos)  # para depuración

def mostrar_datos(datos):
    texto_salida.delete(1.0, tk.END)
    texto_salida.insert(
        tk.END,
        (
            f"Catalog Number: {datos['Catalog Number']}\n"
            f"HP: {datos['Power (HP)']}\n"
            f"RPM: {datos['Speed (RPM)']}\n"
            f"Phase: {datos['Phase']}\n"
            f"Hz: {datos['Hertz']}\n"
            f"Volts: {datos['Voltage']}\n"
            f"Order Codes: {', '.join(datos['Order Codes'])}"
        ),
    )

# --- INTERFAZ ---
root = tk.Tk()
root.title("Extractor de Datos PDF")
aplicar_colorimetria(root)

btn_excel = ttk.Button(root, text="Seleccionar Excel", command=seleccionar_excel)
btn_excel.pack(pady=5)

btn_config = ttk.Button(root, text="Configurar Columnas", command=configurar_columnas)
btn_config.pack(pady=5)

btn_sap = ttk.Button(root, text="Descargar de SAP", command=ejecutar_sap)
btn_sap.pack(pady=5)

btn_cargar = ttk.Button(root, text="Seleccionar PDF", command=seleccionar_pdf)
btn_cargar.pack(pady=5)

texto_salida = tk.Text(
    root,
    height=10,
    width=80,
    bg=ABB_COLORS["bg"],
    fg=ABB_COLORS["fg"],
)
texto_salida.pack(pady=10)

root.mainloop()

